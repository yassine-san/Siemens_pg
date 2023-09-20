import json

from django.contrib.auth import logout
from django.core import serializers
from django.shortcuts import render, redirect
from django.db import connection
from openpyxl import load_workbook
from django.http import JsonResponse
from .models import Quality
from .models import Srs
from .models import Can24
from .models import Partner, Ccr
from users.models import Account, Account
from django.db.models import Count, F, Func, Value, CharField, Max
from datetime import datetime
from datetime import date
from django.core.paginator import Paginator
from django.db.models import Q
from django.db.models import F


def home(request):
    user = request.user
    if not request.user.is_authenticated:
        return redirect('login')

    user_type = user.user_type
    if user_type == 3:
        username = user.partenariat
    else:
        username = {1: "manager", 2: "assistant"}.get(user_type, "unknown")

    print(user.partenariat)
    partenaria = get_user_partenariat(user)
    return render(request, "dashboard/home.html", {"username": username,
                                                   "partenariat": partenaria})


def get_user_partenariat(user):
    user_type = user.user_type
    if user_type == 1 or user_type == 2:
        username = "all"
    elif user_type == 3:
        username = Partner.objects.filter(partnername=user.partenariat).first().partnerid
        # username = user.partenariat
    else:
        username = "unknown"
    return username


def quality_interface(request):
    if not request.user.is_authenticated:
        return redirect('login')

    return render(request, "dashboard/data_quality.html")


def SRS_Connectivity(request):
    results = Srs.objects.values('country_region').annotate(
        equipment_serial_number_count=Count('equipment_serial_number')).order_by('country_region')


    return render(request, 'dashboard/SRS.html',{'countries': results})


def africaIb_interface(request):
    if not request.user.is_authenticated:
        return redirect('login')

    partner = get_user_partenariat(request.user)

    max_week = Quality.objects.aggregate(max_week=Max('week'))['max_week']

    user = request.user
    partner = get_user_partenariat(user)

    if(partner != 'all'):
        # Query to count serial numbers for each cstcountry
        results = Quality.objects.filter(
            week=max_week,
            cstcountry__isnull=False,
            servicepartner=partner,
        ).values('cstcountry').annotate(serialnumber_count=Count('serialnumber'))
    else:
        # Query to count serial numbers for each cstcountry
        results = Quality.objects.filter(
            week=max_week,
            cstcountry__isnull=False,  # Ensures cstCountry is not empty
        ).values('cstcountry').annotate(serialnumber_count=Count('serialnumber'))

    return render(request, "dashboard/africa_IB.html", {"partner": partner,"countries": results})


def get_missing_fl_countries(request):
    user = request.user
    partner = get_user_partenariat(user)

    if request.method == "POST":
        filters_data = request.POST.get('filters')
        filters_dict = json.loads(filters_data)
        status_filter = filters_dict['status']
        substatus_filter = filters_dict['substatus']
        week_filter = filters_dict['week']

    if partner == "all":
        missing_fl_countries = Quality.objects.filter(flcountry='').values(
            'id', 'servicepartner', 'materialnumber', 'serialnumber',
            'modality', 'ivkname', 'status', 'substatus', 'flcountry'
        )
    else:
        missing_fl_countries = Quality.objects.filter(
            flcountry='', servicepartner=partner
        ).values(
            'id', 'servicepartner', 'materialnumber', 'serialnumber',
            'modality', 'ivkname', 'status', 'substatus', 'flcountry'
        )

    if status_filter != ['all']:
        missing_fl_countries = missing_fl_countries.filter(status__in=status_filter)
    if substatus_filter != ['all']:
        missing_fl_countries = missing_fl_countries.filter(substatus__in=substatus_filter)
    if week_filter != ['all']:
        missing_fl_countries = missing_fl_countries.filter(week__in=week_filter)

    return JsonResponse(list(missing_fl_countries), safe=False)


def pieCharter(request):
    partner = get_user_partenariat(request.user)
    if(partner != 'all'):
        modalities = Quality.objects.filter(servicepartner=partner, modality__isnull=False).values('modality').distinct().order_by('modality')
        count_total = Quality.objects.filter(servicepartner=partner,serialnumber__isnull=False).values('serialnumber').distinct().count()
        count_by = (Quality.objects.filter(servicepartner=partner,modality__isnull=False).order_by('modality').values('modality').
                    annotate(serialnumber_count=Count('serialnumber', distinct=True)))
    else:
        modalities = Quality.objects.filter(modality__isnull=False).values('modality').distinct().order_by('modality')
        count_total = Quality.objects.filter(serialnumber__isnull=False).values('serialnumber').distinct().count()
        count_by = (Quality.objects.filter(modality__isnull=False).order_by('modality').values('modality').
                    annotate(serialnumber_count=Count('serialnumber', distinct=True)))

    data = {
        'total': count_total,
        'modalities': list(modalities),
        'count': list(count_by),
    }
    return JsonResponse(data)

def fillin_missing_country(request):
    if request.method == "POST":
        idCountry = request.POST.get('id')
        missingCountry = request.POST.get('missingCountry')
        record = Quality.objects.filter(id=idCountry).first()

        if record:
            record.flcountry = missingCountry
            record.save()

        response_data = {'message': 'country not missing anymore'}

        return JsonResponse(response_data)


def fillin_missing_cst(request):
    if request.method == "POST":
        idCst = request.POST.get('id')
        missingCst = request.POST.get('missingCst')
        record = Quality.objects.filter(id=idCst).first()

        if record:
            record.customername = missingCst
            record.save()

        response_data = {'message': 'customer name not missing anymore'}

        return JsonResponse(response_data)


def get_missing_customer_name(request):
    user = request.user
    partner = get_user_partenariat(user)

    if request.method == "POST":
        filters_data = request.POST.get('filters')
        filters_dict = json.loads(filters_data)
        status_filter = filters_dict['status']
        substatus_filter = filters_dict['substatus']
        week_filter = filters_dict['week']

    if partner == "all":
        missing_customer_names = Quality.objects.filter(customername='').values(
            'id', 'servicepartner', 'materialnumber', 'serialnumber',
            'modality', 'ivkname', 'status', 'substatus', 'customername'
        )
    else:
        missing_customer_names = Quality.objects.filter(
            customername='', servicepartner=partner
        ).values(
            'id', 'servicepartner', 'materialnumber', 'serialnumber',
            'modality', 'ivkname', 'status', 'substatus', 'customername'
        )

    # Apply your filtering logic using the status, substatus, and week filters
    if status_filter != ['all']:
        missing_customer_names = missing_customer_names.filter(status__in=status_filter)
    if substatus_filter != ['all']:
        missing_customer_names = missing_customer_names.filter(substatus__in=substatus_filter)
    if week_filter != ['all']:
        missing_customer_names = missing_customer_names.filter(week__in=week_filter)

    return JsonResponse(list(missing_customer_names), safe=False)


def registerPartenariat(request):
    if not request.user.is_authenticated:
        return redirect('login')

    # serializers.serialize("json", Deliveries.objects.all())
    partners = Partner.objects.distinct().order_by('partnername')
    partners_list_se = serializers.serialize("json", Partner.objects.distinct().order_by('partnername'))
    return render(request, "dashboard/register_user.html",
                  {"partenariats": partners, "partenariats_se": partners_list_se})


def registerNewUser(request):
    if request.method == 'POST':
        username = request.POST["email"]
        password = username + "?"
        partenariat = request.POST["partner"]
        email = username + Partner.objects.filter(partnername=partenariat).values('email').first().get('email')

        Account.objects.create_user(email, password, partenariat, 3)

        response_data = {'message': 'User registered successfully'}

        return JsonResponse(response_data)


def ccr_interface(request):
    if not request.user.is_authenticated:
        return redirect('login')
    partenaria = get_user_partenariat(request.user);
    return render(request, 'dashboard/ccr.html', {"partenariat":partenaria})



def update_ccr_data(request):
    user = request.user
    partenariat = get_user_partenariat(user)
    if request.method == 'POST':
        partner_number = request.POST.get('partner_number')
        modality = request.POST.get('modality')
        country = request.POST.get('country')
        search_query = request.POST.get('searchQuery')
        # apply filters
        queryset = Ccr.objects.all().filter(contract_end_date__gt =date.today())
        queryset_quality = Quality.objects.all()

        if partenariat != 'all' or partner_number :
            queryset = queryset.filter(service_partner_id=partner_number) if partenariat == 'all'  else queryset.filter(service_partner_id=partenariat)

            queryset_quality = queryset_quality.filter(servicepartner=partner_number) if partenariat == 'all'  else queryset_quality.filter(servicepartner=partenariat)    
        if modality :
            queryset = queryset.filter(modality=modality)
            queryset_quality = queryset_quality.filter(modality=modality)
        if country :
            queryset = queryset.filter(country=country)
            queryset_quality = queryset_quality.filter(flcountry=country)
        if search_query :
            queryset = queryset.filter(system_serial_number__contains=search_query)
            queryset_quality = queryset_quality.filter(serialnumber__contains=search_query)

        nb_contract = queryset.filter(system_serial_number__isnull=False,  system_material_number__gt='').count()

        active_systems = queryset_quality.filter(status='active',substatus='active',week = 'Week 30').values('serialnumber').distinct().count()
        
        if active_systems > 0:
            ccr_percent = (nb_contract / active_systems) * 100 if nb_contract > 0 else 0;
        else:
            ccr_percent = 0
        # generate table data html 
        table_data = ''
        if queryset.count() == 0:
            table_data = f"<tr class='' style='background-color:transparent;'>                        <td class='ms-countries-td' style='border: none; font-size:18px;' colspan='11'>No data available</td></tr>"
        else:
            for item in queryset:
                table_data += f"<tr class='ms-countries-tr'>\
                                        <td class='ms-countries-td'>{item.service_partner}</td>\
                                        <td class='ms-countries-td'>{item.service_partner_id}</td>\
                                        <td class='ms-countries-td'>{item.system_material_number}</td>\
                                        <td class='ms-countries-td'>{item.system_serial_number}</td>\
                                        <td class='ms-countries-td'>{item.country}</td>\
                                        <td class='ms-countries-td'>{item.modality}</td>\
                                        <td class='ms-countries-td'>{item.product_name}</td>\
                                        <td class='ms-countries-td'>{item.delivery_date}</td>\
                                        <td class='ms-countries-td'>{item.contract_start_date}</td>\
                                        <td class='ms-countries-td'>{item.contract_end_date}</td>\
                                        <td class='ms-countries-td'>{item.end_customer}</td></tr>"
        return JsonResponse({'table_data': table_data,
                             'nb_contracts': nb_contract,
                             'active_systems': active_systems,
                             'ccr_percent': ccr_percent
                             })


def add_ccr_data(request):
    if request.method == 'POST':
        system_serial_number = request.POST.get('systemSerialNumber')
        system_material_number = request.POST.get('systemMaterialNumber')
        product_name = request.POST.get('productName')
        delivery_date = request.POST.get('deliveryDate')
        handover_date = request.POST.get('handoverDate')
        contract_start_date = request.POST.get('contractStartDate')
        contract_end_date = request.POST.get('contractEndDate')
        contract_number = request.POST.get('contractNumber')
        eos = request.POST.get('eos')
        eod = request.POST.get('eod')
        end_customer = request.POST.get('endCustomer')
        city = request.POST.get('city')
        country = request.POST.get('country')
        modality = request.POST.get('modality')
        service_partner = request.POST.get('servicePartner')
        service_partner_id = request.POST.get('servicePartnerId')

        # Save the form data to the database
        form_data = Ccr.objects.create(
            system_serial_number=system_serial_number,
            system_material_number=system_material_number,
            product_name=product_name,
            delivery_date=delivery_date,
            handover_date=handover_date,
            contract_start_date=contract_start_date,
            contract_end_date=contract_end_date,
            contract_number=contract_number,
            eos=eos,
            eod=eod,
            end_customer=end_customer,
            city=city,
            country=country,
            modality=modality,
            service_partner=service_partner,
            service_partner_id=service_partner_id
        )
    form_data.save()
    response_data = {
        'status': 'success',
        'message': 'Form data saved successfully.'
    }

    return JsonResponse(response_data)


def data_quality(request):
    partner = get_user_partenariat(request.user)

    def calculate_percentage(status, substatus, column_name):
        filter_kwargs = {
            f"{column_name}__regex": r'.+',
        }
        if partner == "all":
            count = Quality.objects.filter(status=status, substatus=substatus, **filter_kwargs).count()
            records = Quality.objects.filter(status=status, substatus=substatus).count()
        else:
            count = Quality.objects.all().filter(servicepartner=partner, status=status, substatus=substatus, **filter_kwargs).count()
            records = Quality.objects.all().filter(servicepartner=partner, status=status, substatus=substatus).count()
        
        records = records if records > 0 else 1
        percentage = (count / records) * 100

        return percentage

    active_active_flcountry_percentage = calculate_percentage('active', 'active', 'flcountry')
    active_active_cstname_percentage = calculate_percentage('active', 'active', 'cstname')
    active_shipped_flcountry_percentage = calculate_percentage('active', 'shipped', 'flcountry')
    active_shipped_cstname_percentage = calculate_percentage('active', 'shipped', 'cstname')
    inactive_on_stock_flcountry_percentage = calculate_percentage('inactive', 'on stock', 'flcountry')
    inactive_on_stock_cstname_percentage = calculate_percentage('inactive', 'on stock', 'cstname')

    weeks = Quality.objects.values('week').distinct().order_by('week')

    context = {
        'active_active_flcountry_percentage': active_active_flcountry_percentage,
        'active_active_cstname_percentage': active_active_cstname_percentage,
        'active_shipped_flcountry_percentage': active_shipped_flcountry_percentage,
        'active_shipped_cstname_percentage': active_shipped_cstname_percentage,
        'inactive_on_stock_flcountry_percentage': inactive_on_stock_flcountry_percentage,
        'inactive_on_stock_cstname_percentage': inactive_on_stock_cstname_percentage,
        'weeks_filter': weeks,
    }

    return render(request, 'dashboard/data_quality.html', context)


def update_dataAjax(request):
    partner = get_user_partenariat(request.user)

    if request.method == "POST":
        filters_data = request.POST.get('filters')
        filters_dict = json.loads(filters_data)
        week_filter = filters_dict['week']

    def calculate_percentage(status, substatus, column_name):
        filter_kwargs = {
            f"{column_name}__regex": r'.+',
        }
        if week_filter != ['all']:
            filter_kwargs['week__in'] = week_filter

        if partner == "all":
            count = Quality.objects.filter(status=status, substatus=substatus, **filter_kwargs).count()
            if week_filter != ['all']:
                records = Quality.objects.filter(status=status, substatus=substatus, week__in=week_filter).count()
            else:
                records = Quality.objects.filter(status=status, substatus=substatus).count()
        else:
            count = Quality.objects.all().filter(servicepartner=partner, status=status, substatus=substatus, **filter_kwargs).count()
            if week_filter == ['all']:
                records = Quality.objects.all().filter(servicepartner=partner, status=status, substatus=substatus).count()
            else:
                records = Quality.objects.filter(servicepartner=partner, status=status, substatus=substatus, week__in=week_filter).count()

        records = records if records > 0 else 1
        percentage = (count / records) * 100

        return percentage
    

    active_active_flcountry_percentage = calculate_percentage('active', 'active', 'flcountry')
    active_active_cstname_percentage = calculate_percentage('active', 'active', 'cstname')
    active_shipped_flcountry_percentage = calculate_percentage('active', 'shipped', 'flcountry')
    active_shipped_cstname_percentage = calculate_percentage('active', 'shipped', 'cstname')
    inactive_on_stock_flcountry_percentage = calculate_percentage('inactive', 'on stock', 'flcountry')
    inactive_on_stock_cstname_percentage = calculate_percentage('inactive', 'on stock', 'cstname')

    updated_data = {
        'active_active_flcountry_percentage': active_active_flcountry_percentage,
        'active_active_cstname_percentage': active_active_cstname_percentage,
        'active_shipped_flcountry_percentage': active_shipped_flcountry_percentage,
        'active_shipped_cstname_percentage': active_shipped_cstname_percentage,
        'inactive_on_stock_flcountry_percentage': inactive_on_stock_flcountry_percentage,
        'inactive_on_stock_cstname_percentage': inactive_on_stock_cstname_percentage,
    }

    return JsonResponse(updated_data)


def logout_user(request):
    message = "error"
    if request.user.is_authenticated:
        logout(request)
        message = "success"

    response_data = {'message': message}
    return JsonResponse(response_data)


def excelImporter(request, file_index=None):
    if not request.user.is_authenticated:
        return redirect('login')

    if file_index == 1:
        template_name = "dashboard/excel_import1.html"
    elif file_index == 2:
        template_name = "dashboard/excel_import2.html"
    else:
        # Handle invalid file_index value
        return redirect('home')  # Redirect to the appropriate page

    return render(request, template_name, {'file_index': file_index})


BATCH_SIZE = 10000


def upload_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        uploaded_file = request.FILES['excel_file']

        if connection is None:
            response_data = {'message': 'Invalid request'}
            JsonResponse(response_data)


        cursor = connection.cursor()

        try:
            workbook = load_workbook(uploaded_file, read_only=True)
            first_sheet = workbook.worksheets[0]
            # Skip the header row
            rows = first_sheet.iter_rows(min_row=2, values_only=True)

            batch = []

            for row in rows:
                batch.append(row)
                if len(batch) >= BATCH_SIZE:
                    insert_query = "INSERT INTO exceltable (week, servicepartnername, division, modality, materialnumber, serialnumber, ivkname, status, customername, cinumber, customerid, ponumber, ordernumber, funclocationcode, eod, eos, citycustomer, cstcity, cstcountry, cstname, cststreet, customerenddate, customernumber, customerstartdate, deliverydate, ddate, flcity, flcomments, flcountry, flstreet, flzip, funclocationname1, funclocationname2, handoverdate, servicepartner, substatus, onstockdetails) VALUES (%s" + (
                            ", %s" * (len(row) - 1)) + ")"
                    cursor.executemany(insert_query, batch)
                    batch = []

            if batch:
                insert_query = "INSERT INTO exceltable (week, servicepartnername, division, modality, materialnumber, serialnumber, ivkname, status, customername, cinumber, customerid, ponumber, ordernumber, funclocationcode, eod, eos, citycustomer, cstcity, cstcountry, cstname, cststreet, customerenddate, customernumber, customerstartdate, deliverydate, ddate, flcity, flcomments, flcountry, flstreet, flzip, funclocationname1, funclocationname2, handoverdate, servicepartner, substatus, onstockdetails) VALUES (%s" + (
                        ", %s" * (len(row) - 1)) + ")"
                cursor.executemany(insert_query, batch)

            connection.commit()
            cursor.close()
            connection.close()
            workbook.close()

            response_data = {'status': 1, 'message': 'Database updated', 'data': None}
        except Exception as e:
            print(e)
            response_data = {'status': -1, 'message': 'Error while processing'}

    else:
        response_data = {'status': -1, 'message': 'Invalid request'}

    return JsonResponse(response_data)


########################################################
# New function for uploading Excel file with two sheets
def upload_another_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        uploaded_file = request.FILES['excel_file']

        if connection is None:
            response_data = {'message': 'Invalid request'}
            return JsonResponse(response_data)

        cursor = connection.cursor()

        try:
            workbook = load_workbook(uploaded_file, read_only=True)

            for sheet in workbook:
                # Skip the header row
                rows = sheet.iter_rows(min_row=2, values_only=True)

                batch = []

                for row in rows:
                    batch.append(row)

                    if len(batch) >= BATCH_SIZE:
                        if sheet.title == 'SRS Connectivity':
                            # Adjust the insert query and table name for the SRS_Connectivity sheet
                            insert_query = "INSERT INTO SRS_Connectivity VALUES (%s" + (", %s" * (len(row) - 1)) + ")"
                        elif sheet.title == 'CAN24':
                            # Adjust the insert query and table name for the CAN24 sheet
                            insert_query = "INSERT INTO CAN24 VALUES (%s" + (", %s" * (len(row) - 1)) + ")"
                        else:
                            # Handle other sheets if needed
                            continue

                        cursor.executemany(insert_query, batch)
                        connection.commit()
                        batch = []

                if batch:
                    if sheet.title == 'SRS Connectivity':
                        insert_query = "INSERT INTO SRS_Connectivity(Date,Equipment_material_number,	Equipment_serial_number	,Equipment_type,	Material_IVK_name,	Material_division_ID,	Material_division_text,	Func_Location_Name, 	Equipment_service_partner_ID, Equipment_service_partner_text, Country_Region, Connection_Status,	Capability,	SRS_Connectivity,	RUH_Readiness,	Data_Sent,	Connection_score) VALUES (%s" + (
                                    ", %s" * (len(row) - 1)) + ")"
                    elif sheet.title == 'CAN24':
                        insert_query = "INSERT INTO CAN24(date,equipment_material_number,serial_number,equipment_serial_number,equipment_type,material_ivk_name,material_division_id,material_division_text,func_location_name,equipment_service_partner_id,equipment_service_partner_text,srs_equipment_configurational_state_text,country_region,connection_status,capability,srs_connectivity,ruh_readiness,data_sent,can24,can24_connectable_per_system_type,can24_connection_per_system_type,can24_data_sent_1,can24_data_sent,can24_data_sent_formatiert,connected_can24_modul,connection_score) VALUES (%s" + (
                                    ", %s" * (len(row) - 1)) + ")"
                    else:
                        # Handle other sheets if needed
                        continue

                    cursor.executemany(insert_query, batch)
                    connection.commit()

            cursor.close()
            connection.close()
            workbook.close()

            response_data = {'status': 1, 'message': 'Another Excel data updated', 'data': None}
        except Exception as e:
            print(e)
            response_data = {'status': -1, 'message': 'Erreur lors l\'execution'}

    else:
        response_data = {'status': -1, 'message': 'Invalid request'}

    return JsonResponse(response_data)


# there, there bi

def getFilterColumnNameSrs(condition):
    column_mapping = {
        'Connected': 'srs_connectivity',
        'X_Connectable': 'srs_connectivity',
        'RUH Ready': 'ruh_readiness',
        'X_not RUH ready': 'ruh_readiness',
        'Data Sent': 'data_sent',
        'X_Data not sent': 'data_sent',
        'Connection active': 'connection_score',
        'Connection not active': 'connection_score',
    }
    column_name = column_mapping.get(condition, None)
    if column_name:
        filter_condition = {column_name: condition}
        return filter_condition
    else:
        # Handle the case when posted_variable doesn't match any of the possibilities
        return None


def get_srs_connectivity_chart_data(request):
    if request.method == "POST":
        dates = Srs.objects.values_list('date', flat=True).distinct().order_by('date')

        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'Connected_counts': [],
            'X_Connectable_counts': [],
        }

        for d in dates:
            connected_count = 0
            x_connectable_count = 0
            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnNameSrs(filter_state)
                    if list(filter_condition.keys())[0] == "srs_connectivity":
                        connected_count = Srs.objects.filter(date=d, srs_connectivity='Connected')
                        x_connectable_count = Srs.objects.filter(date=d, srs_connectivity='X_Connectable')
                    else:
                        connected_count = Srs.objects.filter(date=d, srs_connectivity='Connected',
                                                             **filter_condition)
                        x_connectable_count = Srs.objects.filter(date=d, srs_connectivity='X_Connectable',
                                                                 **filter_condition)
            else:
                connected_count = Srs.objects.filter(date=d, srs_connectivity='Connected')
                x_connectable_count = Srs.objects.filter(date=d, srs_connectivity='X_Connectable')

            try:
                user = request.user
                partner = get_user_partenariat(user)
                if partner != 'all':
                    connected_count = connected_count.annotate(
                            equipment_service_partner_id_no_zeros=Func(
                                F('equipment_service_partner_id'),
                                Value('0'),
                                function='LTRIM',
                                output_field=CharField()
                            )).filter(equipment_service_partner_id_no_zeros=partner)
                    x_connectable_count = x_connectable_count.annotate(
                            equipment_service_partner_id_no_zeros=Func(
                                F('equipment_service_partner_id'),
                                Value('0'),
                                function='LTRIM',
                                output_field=CharField()
                            )).filter(equipment_service_partner_id_no_zeros=partner)
                data['Connected_counts'].append(connected_count.count())
                data['X_Connectable_counts'].append(x_connectable_count.count())
            except Exception as e:
                data['Connected_counts'].append(0)
                data['X_Connectable_counts'].append(0)

            # data['Connected_counts'].append(connected_count.count())
            # data['X_Connectable_counts'].append(x_connectable_count.count())

        return JsonResponse(data)


def get_ruh_readiness_chart_data(request):
    if request.method == "POST":
        dates = Srs.objects.values_list('date', flat=True).distinct().order_by('date')

        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'RUH Ready_counts': [],
            'X_not RUH ready_counts': [],
        }

        for d in dates:
            RUH_Ready_count = 0
            X_not_RUH_ready_count = 0
            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnNameSrs(filter_state)
                    if list(filter_condition.keys())[0] == 'ruh_readiness':
                        RUH_Ready_count = Srs.objects.filter(date=d, ruh_readiness='RUH Ready')
                        X_not_RUH_ready_count = Srs.objects.filter(date=d, ruh_readiness='X_not RUH ready')
                    else:
                        RUH_Ready_count = Srs.objects.filter(date=d, ruh_readiness='RUH Ready',
                                                             **filter_condition)
                        X_not_RUH_ready_count = Srs.objects.filter(date=d, ruh_readiness='X_not RUH ready',
                                                                   **filter_condition)
            else:
                RUH_Ready_count = Srs.objects.filter(date=d, ruh_readiness='RUH Ready')
                X_not_RUH_ready_count = Srs.objects.filter(date=d, ruh_readiness='X_not RUH ready')

            try:
                user = request.user
                partner = get_user_partenariat(user)
                if partner != 'all':
                    RUH_Ready_count = RUH_Ready_count.annotate(
                            equipment_service_partner_id_no_zeros=Func(
                                F('equipment_service_partner_id'),
                                Value('0'),
                                function='LTRIM',
                                output_field=CharField()
                            )).filter(equipment_service_partner_id_no_zeros=partner)
                    X_not_RUH_ready_count = X_not_RUH_ready_count.annotate(
                            equipment_service_partner_id_no_zeros=Func(
                                F('equipment_service_partner_id'),
                                Value('0'),
                                function='LTRIM',
                                output_field=CharField()
                            )).filter(equipment_service_partner_id_no_zeros=partner)
                data['RUH Ready_counts'].append(RUH_Ready_count.count())
                data['X_not RUH ready_counts'].append(X_not_RUH_ready_count.count())
            except Exception as e:
                data['RUH Ready_counts'].append(0)
                data['X_not RUH ready_counts'].append(0)

            # data['RUH Ready_counts'].append(RUH_Ready_count)
            # data['X_not RUH ready_counts'].append(X_not_RUH_ready_count)

        return JsonResponse(data)


def get_Data_Sent_chart_data(request):
    if request.method == "POST":
        dates = Srs.objects.values_list('date', flat=True).distinct().order_by('date')

        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'Data Sent_counts': [],
            'X_Data not sent_counts': [],
        }

        for d in dates:
            Data_Sent_count = 0
            X_Data_not_sent_count = 0
            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnNameSrs(filter_state)
                    if list(filter_condition.keys())[0] == "data_sent":
                        Data_Sent_count = Srs.objects.filter(date=d, data_sent='Data Sent')
                        X_Data_not_sent_count = Srs.objects.filter(date=d, data_sent='X_Data not sent')
                    else:
                        Data_Sent_count = Srs.objects.filter(date=d, data_sent='Data Sent', **filter_condition)
                        X_Data_not_sent_count = Srs.objects.filter(date=d, data_sent='X_Data not sent',
                                                                   **filter_condition)
            else:
                Data_Sent_count = Srs.objects.filter(date=d, data_sent='Data Sent')
                X_Data_not_sent_count = Srs.objects.filter(date=d, data_sent='X_Data not sent')

            try:
                user = request.user
                partner = get_user_partenariat(user)
                if partner != 'all':
                    Data_Sent_count = Data_Sent_count.annotate(
                            equipment_service_partner_id_no_zeros=Func(
                                F('equipment_service_partner_id'),
                                Value('0'),
                                function='LTRIM',
                                output_field=CharField()
                            )).filter(equipment_service_partner_id_no_zeros=partner)
                    X_Data_not_sent_count = X_Data_not_sent_count.annotate(
                            equipment_service_partner_id_no_zeros=Func(
                                F('equipment_service_partner_id'),
                                Value('0'),
                                function='LTRIM',
                                output_field=CharField()
                            )).filter(equipment_service_partner_id_no_zeros=partner)
                data['Data Sent_counts'].append(Data_Sent_count.count())
                data['X_Data not sent_counts'].append(X_Data_not_sent_count.count())
            except Exception as e:
                data['Data Sent_counts'].append(0)
                data['X_Data not sent_counts'].append(0)

            # data['Data Sent_counts'].append(Data_Sent_count)
            # data['X_Data not sent_counts'].append(X_Data_not_sent_count)

        return JsonResponse(data)


def get_Connection_score_chart_data(request):
    if request.method == "POST":
        dates = Srs.objects.values_list('date', flat=True).distinct().order_by('date')

        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'Connection active_counts': [],
            'Connection not active_counts': [],
        }

        for d in dates:
            Connection_active_count = 0
            Connection_not_active_count = 0
            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnNameSrs(filter_state)
                    if list(filter_condition.keys())[0] == "connection_score":
                        Connection_active_count = Srs.objects.filter(date=d,
                                                                     connection_score='Connection active').count()
                        Connection_not_active_count = Srs.objects.filter(date=d,
                                                                         connection_score='Connection not active').count()
                    else:
                        Connection_active_count = Srs.objects.filter(date=d, connection_score='Connection active',
                                                                     **filter_condition).count()
                        Connection_not_active_count = Srs.objects.filter(date=d,
                                                                         connection_score='Connection not active',
                                                                         **filter_condition).count()
            else:
                Connection_active_count = Srs.objects.filter(date=d, connection_score='Connection active').count()
                Connection_not_active_count = Srs.objects.filter(date=d,
                                                                 connection_score='Connection not active').count()
            data['Connection active_counts'].append(Connection_active_count)
            data['Connection not active_counts'].append(Connection_not_active_count)

        return JsonResponse(data)


def CAN24(request):
    # chart_data = get_srs_chart_data()

    results = Srs.objects.values('country_region').annotate(
        equipment_serial_number_count=Count('equipment_serial_number')).order_by('country_region')

    return render(request, 'dashboard/MR_CAN24.html', {'countries': results})


def getFilterColumnName(condition):
    column_mapping = {
        'Connected': 'srs_connectivity',
        'X_Connectable': 'srs_connectivity',
        'RUH Ready': 'ruh_readiness',
        'X_not RUH ready': 'ruh_readiness',
        'Data Sent': 'data_sent',
        'X_Data not sent': 'data_sent',
        'Connection active': 'connection_score',
        'Connection not active': 'connection_score',
        'CAN24 connectable': 'can24_connection_per_system_type',
        'x_Not CAN24 connectable': 'can24_connection_per_system_type',
        'Data sent2': 'can24_data_sent_formatiert',
        'x_Data not sent2': 'can24_data_sent_formatiert',
        'x_Not connectable2': 'can24_data_sent_formatiert',
        'Connected CAN24 Modul': 'connected_can24_modul',
        'x_Not Connectable': 'connected_can24_modul',
        'Not Connected CAN24 Modul': 'connected_can24_modul',
    }
    column_name = column_mapping.get(condition, None)
    if column_name:
        # Create a filter condition using the field name and posted value
        if condition == "Data Sent2" or condition == "X_Data not sent2" or condition == 'x_Data not sent2':
            condition = str(condition).replace("2", "")
        filter_condition = {column_name: condition}
        return filter_condition
    else:
        # Handle the case when posted_variable doesn't match any of the possibilities
        return None


def get_srs_connectivity_chart_data2(request):
    if request.method == "POST":
        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        dates = Can24.objects.values_list('date', flat=True).distinct().order_by('date')
        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'Connected_counts': [],
            'X_Connectable_counts': [],
        }

        for d in dates:
            connected_count = 0
            x_connectable_count = 0
            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnName(filter_state)
                    if list(filter_condition.keys())[0] == "srs_connectivity":
                        connected_count = Can24.objects.filter(date=d, srs_connectivity='Connected').count()
                        x_connectable_count = Can24.objects.filter(date=d, srs_connectivity='X_Connectable').count()
                    else:
                        connected_count = Can24.objects.filter(date=d, srs_connectivity='Connected',
                                                               **filter_condition).count()
                        x_connectable_count = Can24.objects.filter(date=d, srs_connectivity='X_Connectable',
                                                                   **filter_condition).count()
            else:
                connected_count = Can24.objects.filter(date=d, srs_connectivity='Connected').count()
                x_connectable_count = Can24.objects.filter(date=d, srs_connectivity='X_Connectable').count()
            data['Connected_counts'].append(connected_count)
            data['X_Connectable_counts'].append(x_connectable_count)

        return JsonResponse(data)


def get_ruh_readiness_chart_data2(request):
    if request.method == "POST":
        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        dates = Can24.objects.values_list('date', flat=True).distinct().order_by('date')

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'RUH Ready_counts': [],
            'X_not RUH ready_counts': [],
        }

        for d in dates:
            RUH_Ready_count = 0
            X_not_RUH_ready_count = 0
            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnName(filter_state)
                    if list(filter_condition.keys())[0] == "ruh_readiness":
                        RUH_Ready_count = Can24.objects.filter(date=d, ruh_readiness='RUH Ready').count()
                        X_not_RUH_ready_count = Can24.objects.filter(date=d, ruh_readiness='X_not RUH ready').count()
                    else:
                        RUH_Ready_count = Can24.objects.filter(date=d, ruh_readiness='RUH Ready',
                                                               **filter_condition).count()
                        X_not_RUH_ready_count = Can24.objects.filter(date=d, ruh_readiness='X_not RUH ready',
                                                                     **filter_condition).count()
            else:
                RUH_Ready_count = Can24.objects.filter(date=d, ruh_readiness='RUH Ready').count()
                X_not_RUH_ready_count = Can24.objects.filter(date=d, ruh_readiness='X_not RUH ready').count()
            data['RUH Ready_counts'].append(RUH_Ready_count)
            data['X_not RUH ready_counts'].append(X_not_RUH_ready_count)

        return JsonResponse(data)


def get_Data_Sent_chart_data2(request):
    if request.method == "POST":
        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        dates = Can24.objects.values_list('date', flat=True).distinct().order_by('date')

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'Data Sent_counts': [],
            'X_Data not sent_counts': [],
        }

        for d in dates:
            Data_Sent_count = 0
            X_Data_not_sent_count = 0

            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnName(filter_state)
                    if list(filter_condition.keys())[0] == "data_sent":
                        Data_Sent_count = Can24.objects.filter(date=d, data_sent='Data Sent').count()
                        X_Data_not_sent_count = Can24.objects.filter(date=d, data_sent='X_Data not sent').count()
                    else:
                        Data_Sent_count = Can24.objects.filter(date=d, data_sent='Data Sent',
                                                               **filter_condition).count()
                        X_Data_not_sent_count = Can24.objects.filter(date=d, data_sent='X_Data not sent',
                                                                     **filter_condition).count()
            else:
                Data_Sent_count = Can24.objects.filter(date=d, data_sent='Data Sent').count()
                X_Data_not_sent_count = Can24.objects.filter(date=d, data_sent='X_Data not sent').count()
            data['Data Sent_counts'].append(Data_Sent_count)
            data['X_Data not sent_counts'].append(X_Data_not_sent_count)

        return JsonResponse(data)


def get_Connection_score_chart_data2(request):
    if request.method == "POST":

        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        dates = Can24.objects.values_list('date', flat=True).distinct().order_by('date')

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'Connection active_counts': [],
            'Connection not active_counts': [],
        }

        for d in dates:
            Connection_active_count = 0
            Connection_not_active_count = 0

            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnName(filter_state)
                    if list(filter_condition.keys())[0] == "connection_score":
                        Connection_active_count = Can24.objects.filter(date=d,
                                                                       connection_score='Connection active').count()
                        Connection_not_active_count = Can24.objects.filter(date=d,
                                                                           connection_score='Connection not active').count()
                    else:
                        Connection_active_count = Can24.objects.filter(date=d, connection_score='Connection active',
                                                                       **filter_condition).count()
                        Connection_not_active_count = Can24.objects.filter(date=d,
                                                                           connection_score='Connection not active',
                                                                           **filter_condition).count()
            else:
                Connection_active_count = Can24.objects.filter(date=d, connection_score='Connection active').count()
                Connection_not_active_count = Can24.objects.filter(date=d,
                                                                   connection_score='Connection not active').count()
            data['Connection active_counts'].append(Connection_active_count)
            data['Connection not active_counts'].append(Connection_not_active_count)

        return JsonResponse(data)


def get_Can24_Connectable_Systems_chart_data(request):
    if request.method == "POST":

        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')



        dates = Can24.objects.values_list('date', flat=True).distinct().order_by('date')

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'CAN24 connectable_counts': [],
            'x_Not CAN24 connectable_counts': [],
        }

        for d in dates:
            CAN24_connectable_count = 0
            x_Not_CAN24_connectable_count = 0

            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnName(filter_state)
                    print(filter_condition)
                    if list(filter_condition.keys())[0] == "can24_connection_per_system_type":
                        CAN24_connectable_count = Can24.objects.filter(date=d,
                                                                       can24_connection_per_system_type='CAN24 connectable').count()
                        x_Not_CAN24_connectable_count = Can24.objects.filter(date=d,
                                                                             can24_connection_per_system_type='x_Not CAN24 connectable').count()
                    else:
                        CAN24_connectable_count = Can24.objects.filter(date=d,
                                                                       can24_connection_per_system_type='CAN24 connectable',
                                                                       **filter_condition).count()
                        x_Not_CAN24_connectable_count = Can24.objects.filter(date=d,
                                                                             can24_connection_per_system_type='x_Not CAN24 connectable',
                                                                             **filter_condition).count()
            else:
                CAN24_connectable_count = Can24.objects.filter(date=d,
                                                               can24_connection_per_system_type='CAN24 connectable').count()
                x_Not_CAN24_connectable_count = Can24.objects.filter(date=d,
                                                                     can24_connection_per_system_type='x_Not CAN24 connectable').count()

            data['CAN24 connectable_counts'].append(CAN24_connectable_count)
            data['x_Not CAN24 connectable_counts'].append(x_Not_CAN24_connectable_count)

        return JsonResponse(data)


def get_CAN24_Data_Sent_chart_data(request):
    if request.method == "POST":
        postdata = json.loads(request.body.decode('utf-8'))

        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        dates = Can24.objects.values_list('date', flat=True).distinct().order_by('date')
        # dates = [date(2023, 6, 1), date(2023, 6, 15), date(2023, 7, 1), date(2023, 7, 15), date(2023, 8, 1),
        #          date(2023, 8, 15)]

        print(isfilrable)
        print(filter_date)
        print(filter_state)

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'Data sent_counts2': [],
            'x_Data not sent_counts2': [],
            'x_Not connectable_counts2': [],
        }

        for d in dates:
            Data_Sent_count = 0
            X_Data_not_sent_count = 0
            X_Not_connectable = 0
            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnName(filter_state)
                    if list(filter_condition.keys())[0] == "can24_data_sent_formatiert":
                        Data_Sent_count = Can24.objects.filter(date=d, can24_data_sent_formatiert='Data sent').count()
                        X_Data_not_sent_count = Can24.objects.filter(date=d,
                                                                     can24_data_sent_formatiert='x_Data not sent').count()
                        X_Not_connectable = Can24.objects.filter(date=d,
                                                                 can24_data_sent_formatiert='x_Not connectable').count()
                    else:
                        Data_Sent_count = Can24.objects.filter(date=d, can24_data_sent_formatiert='Data sent',
                                                               **filter_condition).count()
                        X_Data_not_sent_count = Can24.objects.filter(date=d,
                                                                     can24_data_sent_formatiert='x_Data not sent',
                                                                     **filter_condition).count()
                        X_Not_connectable = Can24.objects.filter(date=d, can24_data_sent_formatiert='x_Not connectable',
                                                                 **filter_condition).count()
            else:
                Data_Sent_count = Can24.objects.filter(date=d, can24_data_sent_formatiert='Data sent').count()
                X_Data_not_sent_count = Can24.objects.filter(date=d,
                                                             can24_data_sent_formatiert='x_Data not sent').count()
                X_Not_connectable = Can24.objects.filter(date=d, can24_data_sent_formatiert='x_Not connectable').count()
            data['Data sent_counts2'].append(Data_Sent_count)
            data['x_Data not sent_counts2'].append(X_Data_not_sent_count)
            data['x_Not connectable_counts2'].append(X_Not_connectable)

        return JsonResponse(data)


def get_Connected_CAN24_Modul_chart_data(request):
    if request.method == "POST":
        postdata = json.loads(request.body.decode('utf-8'))
        isfilrable = postdata.get('isFiltered')
        filter_date = postdata.get('filter_date')
        filter_state = postdata.get('filter_state')

        dates = Can24.objects.values_list('date', flat=True).distinct().order_by('date')
        # dates = [date(2023, 6, 1), date(2023, 6, 15), date(2023, 7, 1), date(2023, 7, 15), date(2023, 8, 1),
        #          date(2023, 8, 15)]

        data = {
            'labels': [d.strftime('%b %d ,%y') for d in dates],
            'Connected CAN24 Modul_counts': [],
            'x_Not Connectable_counts': [],
            'Not Connected CAN24 Modul_counts': [],
        }

        for d in dates:
            Connected_CAN24_Modul_count = 0
            x_Not_Connectable_count = 0
            Not_Connected_CAN14_Modul_count = 0

            if isfilrable:
                date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
                if d == date_filter:
                    filter_condition = getFilterColumnName(filter_state)
                    if list(filter_condition.keys())[0] == "connected_can24_modul":
                        Connected_CAN24_Modul_count = Can24.objects.filter(date=d,
                                                                           connected_can24_modul='Connected CAN24 Modul').count()
                        x_Not_Connectable_count = Can24.objects.filter(date=d,
                                                                       connected_can24_modul='x_Not Connectable').count()
                        Not_Connected_CAN14_Modul_count = Can24.objects.filter(date=d,
                                                                               connected_can24_modul='Not Connected CAN24 Modul').count()
                    else:
                        Connected_CAN24_Modul_count = Can24.objects.filter(date=d,
                                                                           connected_can24_modul='Connected CAN24 Modul',
                                                                           **filter_condition).count()
                        x_Not_Connectable_count = Can24.objects.filter(date=d,
                                                                       connected_can24_modul='x_Not Connectable',
                                                                       **filter_condition).count()
                        Not_Connected_CAN14_Modul_count = Can24.objects.filter(date=d,
                                                                               connected_can24_modul='Not Connected CAN24 Modul, **filter_condition').count()
            else:
                Connected_CAN24_Modul_count = Can24.objects.filter(date=d,
                                                                   connected_can24_modul='Connected CAN24 Modul').count()
                x_Not_Connectable_count = Can24.objects.filter(date=d,
                                                               connected_can24_modul='x_Not Connectable').count()
                Not_Connected_CAN14_Modul_count = Can24.objects.filter(date=d,
                                                                       connected_can24_modul='Not Connected CAN24 Modul').count()
            data['Connected CAN24 Modul_counts'].append(Connected_CAN24_Modul_count)
            data['x_Not Connectable_counts'].append(x_Not_Connectable_count)
            data['Not Connected CAN24 Modul_counts'].append(Not_Connected_CAN14_Modul_count)

        return JsonResponse(data)


def get_equipment_data(request):
    equipment_data = Srs.objects.all().values(
        'equipment_service_partner_id', 'equipment_service_partner_text', 'country_region',
        'func_location_name', 'equipment_material_number', 'equipment_serial_number',
        'material_division_text', 'material_ivk_name', 'srs_connectivity',
        'ruh_readiness', 'data_sent'
    )

    itemsPerPage = 500  # Adjust this according to your needs
    paginator = Paginator(equipment_data, itemsPerPage)

    page = request.GET.get('page')
    equipment_page = paginator.get_page(page)

    return JsonResponse(list(equipment_page), safe=False)


def get_equipment_dataAjax(request):
    
    if request.method == "POST":
        data = json.loads(request.POST['data'])
        isfilrable = data["isFiltered"]
        filter_date = ''
        filter_state = ''
        if isfilrable:
            filter_date = data["filter_dated"]
            filter_state = data["filter_state"]

        equipment_data = Srs.objects.all()
        max_date =  equipment_data.aggregate(max_date=Max('date'))['max_date']
        count_sys_active = equipment_data.filter(connection_score='Connection active', date=max_date).count()
        user = request.user
        partner = get_user_partenariat(user)

        if partner != 'all':
            print('partner', partner )
            equipment_data = equipment_data.annotate(
                equipment_service_partner_id_no_zeros=Func(
                    F('equipment_service_partner_id'),
                    Value('0'),
                    function='LTRIM',
                    output_field=CharField()
                )
            ).filter(equipment_service_partner_id_no_zeros=partner)
            max_date =  equipment_data.aggregate(max_date=Max('date'))['max_date']
            count_sys_active = Srs.objects.all().filter(connection_score='Connection active',date=max_date).count()
        if isfilrable:
            date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
            filter_condition = getFilterColumnNameSrs(filter_state)
            count_sys_active = Srs.objects.all().filter(connection_score='Connection active',date=date_filter).count();
            equipment_data = equipment_data.filter(date=date_filter, **filter_condition)

        equipment_data = equipment_data.values('equipment_service_partner_id', 'equipment_service_partner_text',
                                               'country_region',
                                               'func_location_name', 'equipment_material_number',
                                               'equipment_serial_number',
                                               'material_division_text', 'material_ivk_name', 'srs_connectivity',
                                               'ruh_readiness', 'data_sent'
                                               )
        
        data = {
            'count_sys_active': count_sys_active,
            'data': list(equipment_data),
        }

        return JsonResponse(data, safe=False)


def get_equipment_data_can24(request):
    equipment_data = Can24.objects.all().values(
        'equipment_service_partner_id', 'equipment_service_partner_text', 'country_region',
        'func_location_name', 'equipment_material_number', 'equipment_serial_number',
        'material_ivk_name', 'srs_connectivity',
        'ruh_readiness', 'data_sent'
    )

    itemsPerPage = 500  # Adjust this according to your needs
    paginator = Paginator(equipment_data, itemsPerPage)

    page = request.GET.get('page')
    equipment_page = paginator.get_page(page)

    return JsonResponse(list(equipment_page), safe=False)


def get_equipment_data_CAN24_Ajax(request):
    if request.method == "POST":
        data = json.loads(request.POST['data'])
        isfilrable = data["isFiltered"]
        filter_date = ''
        filter_state = ''
        if isfilrable:
            filter_date = data["filter_dated"]
            filter_state = data["filter_state"]

        equipment_data = Can24.objects.all()
        max_date =  equipment_data.aggregate(max_date=Max('date'))['max_date']
        count_sys_active = equipment_data.filter(connection_score='Connection active', date = max_date).count()
        user = request.user
        partner = get_user_partenariat(user)
        if partner != 'all':
            print('partner here ', partner)
            partner = "".join(['0000',str(partner)])
            # equipment_data = equipment_data.annotate(
            #     equipment_service_partner_id_no_zeros=Func(
            #         F('equipment_service_partner_id'),
            #         Value('0'),
            #         function='LTRIM',
            #         output_field=CharField()
            #     )
            # ).filter(equipment_service_partner_id_no_zeros=partner)
            #equipment_data = equipment_data.filter(equipment_service_partner_id=partner)
            max_date =  equipment_data.aggregate(max_date=Max('date'))['max_date']
            print("max_date", max_date) 
            count_sys_active = equipment_data.filter(connection_score='Connection active', date=max_date).count()
            print("count_sys_active", count_sys_active)
        if isfilrable:
            print("firtlable")
            date_filter = datetime.strptime(filter_date, '%b %d ,%y').date()
            filter_condition = getFilterColumnName(filter_state)
            count_sys_active = Can24.objects.all().filter(connection_score='Connection active',date=date_filter).count();

            equipment_data = equipment_data.filter(date=date_filter, **filter_condition)

        equipment_data = equipment_data.values(
            'equipment_service_partner_id', 'equipment_service_partner_text', 'country_region',
            'func_location_name', 'equipment_material_number', 'equipment_serial_number',
            'material_ivk_name', 'srs_connectivity',
            'ruh_readiness', 'data_sent'
        )

        data = {
            'count_sys_active': count_sys_active,
            'data': list(equipment_data),
        } 

        return JsonResponse(data, safe=False)




def active_system_count(request):
    # Execute the SQL query
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(*) 
            FROM srs_connectivity 
            WHERE connection_score = 'Connection active'
            AND date = (SELECT MAX(date) FROM srs_connectivity);
        """)
        active_system_count = cursor.fetchone()[0]  # Get the count value

    results = Srs.objects.values('country_region').annotate(
        equipment_serial_number_count=Count('equipment_serial_number')).order_by('country_region')

    # Pass the count value to the template
    return render(request, 'dashboard/SRS.html', {'active_system_count': active_system_count,'countries': results})


def get_equipment_dataAjax2(request):
    partner = get_user_partenariat(request.user)
    equipment_data = Quality.objects.filter(week='Week 30').values(
        'serialnumber', 'materialnumber', 'servicepartnername',
        'servicepartner', 'flcountry', 'status',
        'substatus', 'onstockdetails')
    if partner != 'all':
        equipment_data = equipment_data.filter(servicepartner=partner)

    return JsonResponse(list(equipment_data), safe=False)


def get_all_ivknames(request):
    # Query the Srs table to get all ivknames
    ivknames = Quality.objects.values_list('ivkname', flat=True).distinct().order_by('ivkname')

    # Convert the queryset to a list
    ivkname_list = list(ivknames)

    # Return the ivkname list as JSON response
    return JsonResponse(ivkname_list, safe=False)


# def get_serialnumber_count(request):
#     # Calculate the count of serial numbers
#     serialnumber_count = Quality.objects.filter(week='Week 30').values('serialnumber').distinct().count()

#     return JsonResponse({'serialnumber_count': serialnumber_count})
# def count_on_stock_substatus(request):
#     # Use the annotate function to count the 'on stock' substatus values
#     count = Quality.objects.filter(week='Week 30', substatus='on stock').values('serialnumber').distinct().count()
#     # Create a dictionary with the count result
#     response_data = {'count': count}

#     return JsonResponse(response_data)

# def count_active_status(request):
#     # Use the annotate function to count the 'on stock' substatus values
#     count = Quality.objects.filter(week='Week 30', status='active').values('serialnumber').distinct().count()

#     # Create a dictionary with the count result
#     response_data = {'count': count}

#     return JsonResponse(response_data)

# def count_shipped_substatus(request):
#     # Use the annotate function to count the 'on stock' substatus values
#     count = Quality.objects.filter(week='Week 30', substatus='shipped').values('serialnumber').distinct().count()

#     # Create a dictionary with the count result
#     response_data = {'count': count}

#     return JsonResponse(response_data)
def get_serial_data_by_search(request):
    # Get the search input value from the request
    search_input = request.GET.get('search')

    # Filter the Quality objects by week 30 and serial number containing the search input
    serial_data = Quality.objects.filter(week='Week 30', serialnumber__icontains=search_input).values(
        'serialnumber', 'materialnumber', 'servicepartnername',
        'servicepartner', 'flcountry', 'status',
        'substatus', 'onstockdetails'
    )

    return JsonResponse(list(serial_data), safe=False)


def get_row_data(request):
    # Get the unique identifier from the request
    unique_identifier = request.GET.get('unique_identifier')

    # Query the database to get the row data based on the unique identifier
    try:
        row_data = Quality.objects.get(serialnumber=unique_identifier)
    except Quality.DoesNotExist:
        return JsonResponse({
            'equipments_value': '1',
            'on_stock_value': 'blank',
            'active_value': 'blank',
            'shipped_value': 'blank',
        })

    # Process the row_data to determine card values
    equipments_value = '1'  # Assuming serial number is unique
    on_stock_value = row_data.substatus if row_data.substatus == 'on stock' else 'blank'
    active_value = '1' if row_data.status == 'active' else 'blank'
    shipped_value = row_data.substatus if row_data.substatus == 'shipped' else 'blank'

    # Send the values as JSON response
    return JsonResponse({
        'equipments_value': equipments_value,
        'on_stock_value': on_stock_value,
        'active_value': active_value,
        'shipped_value': shipped_value,
    })


def filter_data(request):
    # Retrieve the selected filter parameters from the request
    modality_filter = request.GET.getlist('modality_filter[]')
    country_filter = request.GET.getlist('country_filter[]')
    service_partner_filter = request.GET.getlist('service_partner_filter[]')
    status_filter = request.GET.getlist('status_filter[]')
    substatus_filter = request.GET.getlist('substatus_filter[]')
    on_stock_details_filter = request.GET.getlist('on_stock_details_filter[]')
    print('Modality Filter:', modality_filter)
    print('Country Filter:', country_filter)

    # Initialize a queryset with all objects from your Quality model
    filtered_data = Quality.objects.filter(week='Week 30')

    # Implement filtering logic based on the selected filters
    if modality_filter:
        filtered_data = filtered_data.filter(modality__in=modality_filter)
        print('modality filter count :', filtered_data.count())

    if country_filter:
        filtered_data = filtered_data.filter(cstcountry__in=country_filter)

    if service_partner_filter:
        filtered_data = filtered_data.filter(servicepartnername__in=service_partner_filter)

    if status_filter:
        filtered_data = filtered_data.filter(status__in=status_filter)

    if substatus_filter:
        filtered_data = filtered_data.filter(substatus__in=substatus_filter)

    if on_stock_details_filter:
        filtered_data = filtered_data.filter(onstockdetails__in=on_stock_details_filter)

    # Serialize the filtered data to JSON
    # You may need to convert the filtered_data queryset to a list or dictionary
    # depending on how you want to structure the JSON response
    filtered_data = {
        'filtered_data': [item.to_dict() for item in filtered_data]  # Sample conversion to a list
    }
    # print('filtered data : ',filtered_data)

    return JsonResponse(filtered_data, safe=False)


# views.py

def get_filtered_counts(request):
    # Retrieve the selected filter parameters from the request
    modality_filter = request.GET.getlist('modality_filter[]')
    print('modality filter count for the cards :', modality_filter)

    country_filter = request.GET.getlist('country_filter[]')
    print('country filter count for the cards :', country_filter)

    service_partner_filter = request.GET.getlist('service_partner_filter[]')
    print('service partnerfilter count for the cards :', service_partner_filter)

    status_filter = request.GET.getlist('status_filter[]')
    print('status filter count for the cards :', status_filter)

    substatus_filter = request.GET.getlist('substatus_filter[]')
    print('substatus filter count for the cards :', substatus_filter)

    on_stock_details_filter = request.GET.getlist('on_stock_details_filter[]')
    print('on stock filter count for the cards :', on_stock_details_filter)

    # Create a filter dictionary based on selected filters
    filter_dict = {}
    partner = get_user_partenariat(request.user)
    if partner != 'all':
        filter_dict['servicepartner'] = partner

    if modality_filter:
        filter_dict['modality__in'] = modality_filter
    if country_filter:
        filter_dict['cstcountry__in'] = country_filter
    if service_partner_filter:
        filter_dict['servicepartner__in'] = service_partner_filter
    if status_filter:
        filter_dict['status__in'] = status_filter
    if substatus_filter:
        filter_dict['substatus__in'] = substatus_filter
    if on_stock_details_filter:
        filter_dict['onstockdetails__in'] = on_stock_details_filter
    print('Filter Dictionary:', filter_dict)

    # Calculate counts based on the selected filters
    counts = {
        'serialnumber_count': Quality.objects.filter(week='Week 30', **filter_dict).values(
            'serialnumber').distinct().count(),
        'on_stock_count': Quality.objects.filter(week='Week 30', **filter_dict, substatus='on stock').values(
            'serialnumber').distinct().count(),
        'active_count': Quality.objects.filter(week='Week 30', **filter_dict, status='active',
                                               substatus='active').values('serialnumber').distinct().count(),
        'shipped_count': Quality.objects.filter(week='Week 30', **filter_dict, status='active',
                                                substatus='shipped').values('serialnumber').distinct().count(),
    }

    return JsonResponse(counts)


def manage_user_interface(request):
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'dashboard/manage_users.html')


def update_manage_users(request):
    if request.method == 'POST':
        users_list = Account.objects.all()
        search_query = request.POST.get('search_query')
        if search_query:
            users_list = users_list.filter(email__icontains=search_query)

        table_data = ''
        for user in users_list:
            partenariat = get_user_partenariat(user)
            active_status = "<span style='border-radius: 5px; font-size: 0.875em;font-weight: 600; background-color:#0ba75b;padding: 3px 20px; '>active</span>"
            inactive_status = "<span style='border-radius: 5px; font-size: 0.875em;font-weight: 600; background-color:red; padding: 3px 20px;'>inactive</span>"
            table_data += f"<tr class='data-tr'>\
                    <td>{user.email}</td>\
                    <td>{user.partenariat}</td>\
                    <td>{ active_status if user.is_active else inactive_status}</td>\
                    <td><button class='delete-btn-user' onclick='delete_user({user.id})'><i class='fa fa-trash' aria-hidden='true'></i></button><button class='reset-btn-user' onclick='reset_user({user.id})'><i class='fa fa-repeat' aria-hidden='true'></i></button> </td>\
                        </tr><tr class='spacer'><td colspan='100'></td></tr>" if partenariat != 'all' else ""
            
        return JsonResponse({'table_data': table_data})
    
def delete_user(request):
    if request.method == 'POST':
        user_id = request.POST['user_id']
        user = Account.objects.get(id=user_id)
        user.delete()
        return JsonResponse({'message': 'User deleted successfully!'})
    
def reset_user(request):
    if request.method == 'POST':
        user_id = request.POST['user_id']
        user = Account.objects.get(id=user_id)
        user.set_password(user.email.split("@")[0]+"?")
        user.is_active = False
        user.save()

        return JsonResponse({'message': 'User password reset successfully!'})
     